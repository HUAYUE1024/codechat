"""
Enhanced CodeAgent v2 — Inspired by Claude Code Architecture

Key improvements from Claude Code:
1. Tool System: Permission checks, concurrency control, progress reporting
2. Memory Management: Context compression, smart truncation
3. Agent Architecture: Multi-agent collaboration support
4. System Prompt: Structured prompt engineering
5. Execution: Timeout, retry, result caching

Architecture:
    ┌─────────────────────────────────────────────────────────────┐
    │                      Agent Orchestrator                      │
    ├─────────────────────────────────────────────────────────────┤
    │  Planning  │  Memory  │  Tools  │  Action  │  Reflection   │
    │    拆解任务  │  上下文管理 │  工具集  │  执行引擎  │    反思优化    │
    └─────────────────────────────────────────────────────────────┘
"""

from __future__ import annotations

import hashlib
import json
import sys
import os
import re
import time
import threading
from abc import ABC, abstractmethod
from collections.abc import Callable
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Protocol, runtime_checkable
import concurrent.futures

from .config import get_snowcode_dir
from .rag import _get_llm_config, _call_llm
from .scanner import scan_files, read_file
from .store import VectorStore


# ═══════════════════════════════════════════════════════════════════════
#  Core Types & Protocols
# ═══════════════════════════════════════════════════════════════════════

class ToolPermission(Enum):
    """Tool permission levels inspired by Claude Code."""
    ALLOWED = "allowed"          # Auto-allowed, no prompt
    PROMPT = "prompt"            # Ask user for permission
    DENIED = "denied"            # Blocked by policy
    DANGEROUS = "dangerous"      # Destructive operation


@dataclass
class ToolResult:
    """Enhanced tool result with metadata."""
    success: bool
    output: str
    tool_name: str
    elapsed_ms: float = 0
    is_truncated: bool = False
    metadata: dict = field(default_factory=dict)
    
    @property
    def preview(self) -> str:
        """Get truncated preview for display."""
        max_len = 200
        if len(self.output) <= max_len:
            return self.output
        return self.output[:max_len] + "..."


@dataclass
class ToolExecutionContext:
    """Context passed to tools during execution."""
    root: Path
    store: VectorStore | None = None
    llm: Any = None
    agent_id: str = ""
    session_id: str = ""
    abort_signal: threading.Event | None = None
    on_progress: Callable[[str], None] | None = None


@runtime_checkable
class ToolProtocol(Protocol):
    """Protocol defining tool interface."""
    name: str
    description: str
    
    @property
    def parameters(self) -> dict[str, str]: ...
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str: ...
    
    def check_permission(self, params: dict) -> ToolPermission:
        """Check if tool execution requires permission."""
        ...
    
    def is_read_only(self) -> bool:
        """Whether tool only reads data."""
        ...
    
    def is_concurrent_safe(self) -> bool:
        """Whether tool can run concurrently with others."""
        ...


# ═══════════════════════════════════════════════════════════════════════
#  Enhanced Tool Base Class
# ═══════════════════════════════════════════════════════════════════════

class BaseTool(ABC):
    """Enhanced base tool with Claude Code features."""
    
    name: str = ""
    description: str = ""
    max_result_size: int = 30000  # Max chars before truncation
    search_hint: str = ""  # For tool discovery
    
    @property
    def parameters(self) -> dict[str, str]:
        return {}
    
    @abstractmethod
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        """Execute the tool."""
        ...
    
    def check_permission(self, params: dict) -> ToolPermission:
        """Default: PROMPT for write operations, ALLOWED for reads."""
        return ToolPermission.PROMPT if not self.is_read_only() else ToolPermission.ALLOWED
    
    def is_read_only(self) -> bool:
        """Override in subclasses."""
        return False
    
    def is_concurrent_safe(self) -> bool:
        """Read-only tools are generally safe to run concurrently."""
        return self.is_read_only()
    
    def validate_input(self, params: dict) -> str | None:
        """Validate input parameters. Returns error message or None."""
        return None
    
    def format_output(self, output: str) -> str:
        """Format and truncate output if needed."""
        if len(output) <= self.max_result_size:
            return output
        return output[:self.max_result_size] + f"\n... [Truncated: {len(output)} chars total]"
    
    def get_activity_description(self, params: dict) -> str:
        """Human-readable description for progress display."""
        return f"Running {self.name}"


# ═══════════════════════════════════════════════════════════════════════
#  Enhanced Tools Implementation
# ═══════════════════════════════════════════════════════════════════════

class SearchTool(BaseTool):
    """Semantic code search with query expansion."""
    
    name = "search"
    description = "语义搜索代码库，返回最相关的代码片段"
    search_hint = "search code semantically"
    
    @property
    def parameters(self):
        return {"query": "搜索关键词", "n": "结果数量(默认5)"}
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        if not ctx.store:
            return "Error: VectorStore not available"
        
        query = params.get("query", "")
        n = int(params.get("n", 5))
        if not query:
            return "Error: query required"
        
        # Query expansion via LLM
        expanded_query = self._expand_query(query, ctx.llm)
        
        results = ctx.store.query(expanded_query, n_results=n)
        if not results:
            return "No results found."
        
        parts = []
        for i, r in enumerate(results, 1):
            m = r["metadata"]
            parts.append(
                f"[{i}] `{m['file_path']}` L{m['start_line']}-{m['end_line']}\n"
                f"```\n{r['content']}\n```"
            )
        return "\n\n".join(parts)
    
    def _expand_query(self, query: str, llm: Any) -> str:
        """Expand query with technical terms."""
        if not llm or not llm.available:
            return query
        
        try:
            prompt = (
                "You are a code search expert. Convert the user question into "
                "rich keywords for vector search. Include possible English variable names, "
                "function names, and technical terms. Return only keywords, no explanation."
            )
            expanded = llm.complete(prompt, query).strip()
            if expanded and len(expanded) < 200:
                return f"{query} {expanded}"
        except Exception:
            pass
        return query
    
    def get_activity_description(self, params: dict) -> str:
        query = params.get("query", "")
        return f"Searching for: {query[:50]}"


class ReadFileTool(BaseTool):
    """Read file with smart truncation."""
    
    name = "read_file"
    description = "读取文件完整内容，默认读整个文件，可选指定行范围"
    search_hint = "read file contents"
    max_lines_per_read = 2000
    
    @property
    def parameters(self):
        return {
            "path": "文件路径",
            "start": "起始行(可选,默认读全部)",
            "end": "结束行(可选)"
        }
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        path = params.get("path", "")
        if not path:
            return "Error: path required"
        
        # Resolve and validate path
        full = (ctx.root / path).resolve()
        if not full.is_relative_to(ctx.root):
            return "Access denied: path outside project root"
        
        if not full.exists():
            # Try glob search
            name = Path(path).name
            matches = [m for m in ctx.root.rglob(name) if m.resolve().is_relative_to(ctx.root)]
            if matches:
                full = matches[0]
            else:
                return f"File not found: {path}"
        
        content = read_file(full)
        if content is None:
            return f"Cannot read: {path}"
        
        lines = content.splitlines()
        total = len(lines)
        s = max(1, int(params.get("start", 1)))
        e = min(total, int(params.get("end", total)))
        
        # Smart truncation
        if e - s + 1 > self.max_lines_per_read:
            e = s + self.max_lines_per_read - 1
            trunc_msg = f"\n... [File has {total} lines, showing {s}-{e}. Use start={e+1} for more]"
        else:
            trunc_msg = ""
        
        numbered = "\n".join(f"{i+s:>4} | {l}" for i, l in enumerate(lines[s-1:e]))
        rel = str(full.relative_to(ctx.root))
        return f"`{rel}` ({total} lines, showing {s}-{e})\n```\n{numbered}{trunc_msg}\n```"
    
    def get_activity_description(self, params: dict) -> str:
        path = params.get("path", "")
        return f"Reading {path}"


class FindPatternTool(BaseTool):
    """Regex pattern search with ReDoS protection."""
    
    name = "find_pattern"
    description = "正则搜索代码（找函数定义、import等）"
    search_hint = "search with regex pattern"
    max_results = 30
    
    @property
    def parameters(self):
        return {"pattern": "正则表达式", "file_glob": "文件过滤(可选)"}
    
    def is_read_only(self) -> bool:
        return True
    
    def validate_input(self, params: dict) -> str | None:
        pattern = params.get("pattern", "")
        if not pattern:
            return "Error: pattern required"
        if len(pattern) > 200:
            return "Error: pattern too long (max 200 chars)"
        
        # ReDoS protection
        if re.search(r'(\([^)]+\)|\w+)([*+]{2,}|[*+]\?)', pattern):
            return "Error: pattern rejected due to potential ReDoS vulnerability"
        
        return None
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        # Validate first
        error = self.validate_input(params)
        if error:
            return error
        
        pattern = params.get("pattern", "")
        file_glob = params.get("file_glob")
        
        try:
            regex = re.compile(pattern)
        except re.error as e:
            return f"Invalid regex: {e}"
        
        files = scan_files(ctx.root)
        all_matches = []
        cancel_event = threading.Event()
        
        def search_file(f: Path) -> list[str]:
            if cancel_event.is_set():
                return []
            matches = []
            if file_glob:
                rel = str(f.relative_to(ctx.root))
                if not Path(rel).match(file_glob):
                    return matches
            
            content = read_file(f)
            if content is None:
                return matches
            
            for i, line in enumerate(content.splitlines(), 1):
                if len(line) > 500:
                    continue
                try:
                    if regex.search(line):
                        rel = str(f.relative_to(ctx.root))
                        matches.append(f"`{rel}:{i}`  {line.strip()}")
                except Exception:
                    continue
            return matches
        
        with concurrent.futures.ThreadPoolExecutor() as executor:
            futures = [executor.submit(search_file, f) for f in files]
            
            for future in concurrent.futures.as_completed(futures):
                matches = future.result()
                if matches:
                    all_matches.extend(matches)
                if len(all_matches) >= self.max_results:
                    cancel_event.set()
                    for f in futures:
                        f.cancel()
                    break
        
        return "\n".join(all_matches[:self.max_results]) if all_matches else "No matches."
    
    def get_activity_description(self, params: dict) -> str:
        pattern = params.get("pattern", "")
        return f"Searching pattern: {pattern[:30]}"


class ListDirTool(BaseTool):
    """List directory structure."""
    
    name = "list_dir"
    description = "列出目录结构"
    search_hint = "list directory contents"
    
    @property
    def parameters(self):
        return {"path": "目录路径(留空=根目录)", "depth": "深度(默认2)"}
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        path = params.get("path", "")
        target = (ctx.root / path).resolve() if path else ctx.root
        
        if not target.is_relative_to(ctx.root):
            return "Access denied: path outside project root"
        
        depth = int(params.get("depth", 2))
        if not target.exists():
            return f"Not found: {path or '.'}"
        
        lines = []
        self._walk(target, depth, lines, "")
        return "\n".join(lines) if lines else "(empty)"
    
    def _walk(self, d: Path, depth: int, lines: list, prefix: str):
        if depth <= 0:
            return
        
        skip = {".git", "__pycache__", "node_modules", ".venv", "venv", ".snowcode", "dist", "build"}
        try:
            entries = sorted(d.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
        except PermissionError:
            return
        
        for e in entries:
            if e.name in skip or e.name.startswith("."):
                continue
            if e.is_dir():
                lines.append(f"{prefix}{e.name}/")
                self._walk(e, depth - 1, lines, prefix + "  ")
            else:
                try:
                    sz = e.stat().st_size
                    ss = f"{sz:,}B" if sz < 1024 else f"{sz // 1024:,}KB"
                except OSError:
                    ss = "?"
                lines.append(f"{prefix}{e.name}  [{ss}]")


class WriteFileTool(BaseTool):
    """Write file with backup."""
    
    name = "write_file"
    description = "写入或覆盖整个文件内容"
    search_hint = "write or create file"
    
    @property
    def parameters(self):
        return {"path": "文件路径", "content": "要写入的完整文件内容"}
    
    def is_read_only(self) -> bool:
        return False
    
    def check_permission(self, params: dict) -> ToolPermission:
        return ToolPermission.ALLOWED
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        path = params.get("path", "")
        content = params.get("content", "")
        
        if not path or not content:
            return "Error: missing path or content arguments"
        
        try:
            safe_path = path.lstrip("/").lstrip("\\")
            full = (ctx.root / safe_path).resolve()
            
            if not full.is_relative_to(ctx.root):
                return f"Error: Cannot write outside project root: {path}"
            
            full.parent.mkdir(parents=True, exist_ok=True)
            
            # Create backup if file exists
            if full.exists():
                import shutil
                backup = full.with_suffix(full.suffix + ".bak")
                shutil.copy2(full, backup)
            
            full.write_text(content, encoding="utf-8")
            return f"Successfully wrote to `{path}` ({len(content)} chars)"
        except Exception as e:
            return f"Error writing file: {e}"
    
    def get_activity_description(self, params: dict) -> str:
        path = params.get("path", "")
        return f"Writing to {path}"


class SearchReplaceTool(BaseTool):
    """Search and replace in file."""
    
    name = "search_replace"
    description = "搜索并替换文件中的指定代码块"
    search_hint = "find and replace in file"
    
    @property
    def parameters(self):
        return {
            "path": "文件路径",
            "old_str": "要被替换的原始代码块(需精确匹配)",
            "new_str": "新的代码块"
        }
    
    def is_read_only(self) -> bool:
        return False
    
    def check_permission(self, params: dict) -> ToolPermission:
        return ToolPermission.ALLOWED
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        path = params.get("path", "")
        old_str = params.get("old_str", "")
        new_str = params.get("new_str", "")
        
        if not path or not old_str:
            return "Error: path and old_str required"
        
        full = (ctx.root / path).resolve()
        if not full.is_relative_to(ctx.root):
            return "Access denied: path outside project root"
        
        if not full.exists():
            return f"Error: File not found `{path}`"
        
        try:
            content = full.read_text(encoding="utf-8")
            if old_str not in content:
                return "Error: old_str not found in the file"
            
            # Create backup
            import shutil
            backup = full.with_suffix(full.suffix + ".bak")
            shutil.copy2(full, backup)
            
            new_content = content.replace(old_str, new_str, 1)
            full.write_text(new_content, encoding="utf-8")
            return f"Successfully replaced code in `{path}`"
        except Exception as e:
            return f"Error modifying file: {e}"


class ShellTool(BaseTool):
    """Execute shell commands with safety checks."""
    
    name = "shell"
    description = "执行终端命令（cmd/bash），如查看git状态、运行测试等"
    search_hint = "execute shell command"
    timeout_seconds = 30
    
    _BLOCKED_COMMANDS = ("rm -rf /", "format", "del /f /s", "shutdown", "mkfs", "dd if=")
    
    @property
    def parameters(self):
        return {"command": "要执行的命令", "cwd": "工作目录(可选,默认项目根目录)"}
    
    def is_read_only(self) -> bool:
        return False
    
    def check_permission(self, params: dict) -> ToolPermission:
        cmd = params.get("command", "").lower()
        for blocked in self._BLOCKED_COMMANDS:
            if blocked in cmd:
                return ToolPermission.DENIED
        return ToolPermission.ALLOWED
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        import subprocess
        
        cmd = params.get("command", "")
        cwd = params.get("cwd", "")
        
        if not cmd:
            return "Error: command required"
        
        # Safety check
        for blocked in self._BLOCKED_COMMANDS:
            if blocked in cmd.lower():
                return f"Error: blocked dangerous command: {cmd}"
        
        work_dir = (ctx.root / cwd).resolve() if cwd else ctx.root
        if not work_dir.is_relative_to(ctx.root) and work_dir != ctx.root:
            return "Error: cwd outside project root"
        
        try:
            result = subprocess.run(
                cmd, shell=True, cwd=str(work_dir),
                capture_output=True, text=True, timeout=self.timeout_seconds,
                encoding="utf-8", errors="replace",
            )
            output = result.stdout + result.stderr
            
            if len(output) > 5000:
                output = output[:5000] + f"\n... [truncated, total {len(output)} chars]"
            
            exit_info = f"[exit code: {result.returncode}]" if result.returncode != 0 else ""
            return f"{output}{exit_info}" if output else f"(no output) {exit_info}"
        except subprocess.TimeoutExpired:
            return f"Error: command timed out ({self.timeout_seconds}s limit)"
        except Exception as e:
            return f"Error: {type(e).__name__}: {e}"
    
    def get_activity_description(self, params: dict) -> str:
        cmd = params.get("command", "")
        return f"Running: {cmd[:50]}"


# ═══════════════════════════════════════════════════════════════════════
#  Multimodal Tools for Agent2
# ═══════════════════════════════════════════════════════════════════════

class _ImageReaderTool(BaseTool):
    """Read images and extract information including OCR text."""
    name = "image_reader"
    description = "读取图片文件，支持OCR提取文字、分析图片内容"
    search_hint = "read image file"
    
    @property
    def parameters(self):
        return {
            "path": "图片文件路径（相对于项目根目录）",
            "mode": "读取模式: 'ocr'(提取文字), 'analyze'(AI分析图片内容), 'info'(基本信息)"
        }
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        file_path = (ctx.root / params["path"]).resolve()
        mode = params.get("mode", "info")
        
        if not file_path.exists():
            return f"Error: 文件不存在: {params['path']}"
        
        supported = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.tif', '.svg'}
        if file_path.suffix.lower() not in supported:
            return f"Error: 不支持的图片格式 '{file_path.suffix}'，支持: {', '.join(supported)}"
        
        info_lines = [f"文件: {file_path.name}", f"路径: {file_path}",
                      f"大小: {self._format_size(file_path.stat().st_size)}"]
        
        try:
            from PIL import Image
            with Image.open(file_path) as img:
                info_lines.extend([f"尺寸: {img.width}x{img.height} 像素", f"模式: {img.mode}"])
        except Exception as e:
            info_lines.append(f"[警告] 读取图片信息失败: {e}")
        
        result = "\n".join(info_lines)
        
        if mode == "ocr":
            result += "\n\n--- OCR 提取的文字 ---\n"
            try:
                import pytesseract
                from PIL import Image
                with Image.open(file_path) as img:
                    text = pytesseract.image_to_string(img, lang='chi_sim+eng')
                    result += text.strip() if text.strip() else "(未检测到文字)"
            except ImportError:
                result += "[警告] pytesseract 未安装。请安装: pip install pytesseract"
            except Exception as e:
                result += f"OCR 失败: {e}"
        elif mode == "analyze":
            result += "\n\n--- AI 图片分析 ---\n"
            if ctx.llm and ctx.llm.available:
                analysis = ctx.llm.analyze_image(
                    system="你是一个图片分析助手。请详细描述这张图片的内容，包括场景、物体、文字、颜色、布局等。",
                    user_text="请描述这张图片",
                    image_path=str(file_path)
                )
                result += analysis if analysis else "[警告] 分析结果为空"
            else:
                result += "[警告] LLM 不可用，无法分析图片"
        
        return result
    
    def _format_size(self, size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"


class _PDFReaderTool(BaseTool):
    """Read and parse PDF documents."""
    name = "pdf_reader"
    description = "读取PDF文档，提取文字内容、元信息和结构"
    search_hint = "read PDF file"
    
    @property
    def parameters(self):
        return {
            "path": "PDF文件路径（相对于项目根目录）",
            "pages": "页码范围，如 '1-5' 或 'all'(默认前10页)"
        }
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        file_path = (ctx.root / params["path"]).resolve()
        pages_param = params.get("pages", "all")
        
        if not file_path.exists():
            return f"Error: 文件不存在: {params['path']}"
        if file_path.suffix.lower() != '.pdf':
            return f"Error: 不是PDF文件: {params['path']}"
        
        result_lines = [f"PDF文档: {file_path.name}", f"路径: {file_path}",
                        f"大小: {self._format_size(file_path.stat().st_size)}", ""]
        
        try:
            import fitz
            doc = fitz.open(str(file_path))
            result_lines.append(f"总页数: {len(doc)}")
            
            pages_to_read = self._parse_page_range(pages_param, len(doc))
            result_lines.append(f"\n--- 文字内容 (第 {pages_to_read[0]+1}-{pages_to_read[-1]+1} 页) ---\n")
            
            total_text = []
            for page_num in pages_to_read:
                page = doc[page_num]
                text = page.get_text()
                if text.strip():
                    total_text.append(f"\n{'='*50}\n第 {page_num + 1} 页\n{'='*50}\n{text.strip()}")
            
            combined = "\n".join(total_text)
            result_lines.append(combined[:10000] + ("\n\n... [已截断]" if len(combined) > 10000 else ""))
            doc.close()
        except ImportError:
            result_lines.append("[警告] PyMuPDF 未安装。请安装: pip install PyMuPDF")
        except Exception as e:
            result_lines.append(f"读取PDF失败: {e}")
        
        return "\n".join(result_lines)
    
    def _parse_page_range(self, pages_param: str, total_pages: int) -> list[int]:
        if pages_param.lower() == "all":
            return list(range(min(total_pages, 10)))
        pages = []
        try:
            for part in pages_param.split(","):
                part = part.strip()
                if "-" in part:
                    start, end = part.split("-", 1)
                    pages.extend(range(max(0, int(start.strip())-1), min(total_pages, int(end.strip()))))
                else:
                    page = int(part.strip()) - 1
                    if 0 <= page < total_pages:
                        pages.append(page)
        except ValueError:
            return list(range(min(total_pages, 10)))
        return sorted(set(pages)) if pages else list(range(min(total_pages, 10)))
    
    def _format_size(self, size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"


class _DocumentReaderTool(BaseTool):
    """Read various document formats (docx, xlsx, csv, txt, md, etc.)."""
    name = "document_reader"
    description = "读取多种格式的文档：Word(.docx)、Excel(.xlsx)、CSV、TXT、Markdown等"
    search_hint = "read document file"
    
    @property
    def parameters(self):
        return {"path": "文件路径（相对于项目根目录）", "max_lines": "最大返回行数(默认500)"}
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        file_path = (ctx.root / params["path"]).resolve()
        max_lines = int(params.get("max_lines", 500))
        
        if not file_path.exists():
            return f"Error: 文件不存在: {params['path']}"
        
        ext = file_path.suffix.lower()
        result_lines = [f"文档: {file_path.name}", f"路径: {file_path}",
                        f"大小: {self._format_size(file_path.stat().st_size)}", f"格式: {ext}", ""]
        
        try:
            if ext == '.docx':
                from docx import Document
                doc = Document(str(file_path))
                result_lines.append("\n".join(p.text for p in doc.paragraphs if p.text.strip())[:10000])
            elif ext == '.xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(str(file_path), data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    result_lines.append(f"\n工作表: {sheet_name}")
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= max_lines:
                            result_lines.append("... [已截断]")
                            break
                        result_lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                wb.close()
            elif ext == '.csv':
                import csv
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    for i, row in enumerate(csv.reader(f)):
                        if i >= max_lines:
                            result_lines.append("... [已截断]")
                            break
                        result_lines.append(" | ".join(row))
            else:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    lines = f.readlines()
                    content = "".join(lines[:max_lines])
                    if len(lines) > max_lines:
                        content += f"\n\n... [已截断，共 {len(lines)} 行]"
                    result_lines.append(content)
        except ImportError as e:
            result_lines.append(f"[警告] 缺少依赖库: {e}")
        except Exception as e:
            result_lines.append(f"读取文件失败: {e}")
        
        return "\n".join(result_lines)
    
    def _format_size(self, size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"


class _FileBrowserTool(BaseTool):
    """Browse files in a local directory."""
    name = "file_browser"
    description = "浏览本地目录中的文件列表，支持按类型筛选"
    search_hint = "browse directory files"
    
    @property
    def parameters(self):
        return {
            "path": "目录路径（相对于项目根目录，默认为项目根目录）",
            "file_type": "文件类型筛选: 'all', 'images', 'docs', 'code', 'data'",
            "recursive": "是否递归搜索子目录 (true/false)"
        }
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        dir_path = (ctx.root / params.get("path", ".")).resolve()
        file_type = params.get("file_type", "all")
        recursive_val = params.get("recursive", False)
        if isinstance(recursive_val, str):
            recursive = recursive_val.lower() == "true"
        else:
            recursive = bool(recursive_val)
        
        if not dir_path.exists() or not dir_path.is_dir():
            return f"Error: 目录不存在: {params.get('path', '.')}"
        if not dir_path.is_relative_to(ctx.root):
            return f"Error: 路径超出项目根目录"
        
        type_map = {
            'images': {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.svg'},
            'docs': {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.txt', '.md'},
            'code': {'.py', '.js', '.ts', '.java', '.c', '.cpp', '.go', '.rs', '.css', '.html'},
            'data': {'.json', '.csv', '.xml', '.yaml', '.yml', '.toml', '.nc', '.nc4'},
        }
        
        files = list(dir_path.rglob("*") if recursive else dir_path.iterdir())
        files = [f for f in files if f.is_file()]
        
        if file_type != "all":
            extensions = type_map.get(file_type, set())
            files = [f for f in files if f.suffix.lower() in extensions]
        
        files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        
        lines = [f"目录: {dir_path.relative_to(ctx.root) if dir_path != ctx.root else '.'}",
                 f"文件数: {len(files)}", f"类型: {file_type}", ""]
        
        if not files:
            lines.append("(没有找到匹配的文件)")
            return "\n".join(lines)
        
        by_ext = {}
        for f in files[:100]:
            ext = f.suffix.lower() or "(无扩展名)"
            by_ext.setdefault(ext, []).append(f)
        
        for ext, ext_files in sorted(by_ext.items()):
            lines.append(f"\n{ext} ({len(ext_files)} 个文件):")
            for f in ext_files[:10]:
                rel = f.relative_to(ctx.root)
                lines.append(f"  {rel}  ({self._format_size(f.stat().st_size)})")
            if len(ext_files) > 10:
                lines.append(f"  ... 及其他 {len(ext_files) - 10} 个文件")
        
        return "\n".join(lines)
    
    def _format_size(self, size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"


class _NCDataReaderTool(BaseTool):
    """Read and analyze NetCDF (.nc) scientific data files."""
    name = "nc_reader"
    description = "读取NetCDF(.nc)科学数据文件，支持查看维度、变量、属性和数据"
    search_hint = "read NetCDF scientific data"
    
    @property
    def parameters(self):
        return {
            "path": "NC文件路径（相对于项目根目录）",
            "mode": "读取模式: 'info'(文件信息), 'vars'(变量列表), 'data'(变量数据), 'stats'(统计信息)",
            "variable": "变量名（mode为data或stats时必需）"
        }
    
    def is_read_only(self) -> bool:
        return True
    
    def run(self, params: dict, ctx: ToolExecutionContext) -> str:
        file_path = (ctx.root / params["path"]).resolve()
        mode = params.get("mode", "info")
        variable = params.get("variable", "")
        
        if not file_path.exists():
            return f"Error: 文件不存在: {params['path']}"
        if file_path.suffix.lower() not in ('.nc', '.nc4', '.netcdf'):
            return f"Error: 不是NetCDF文件: {params['path']}"
        
        try:
            import netCDF4 as nc
            dataset = nc.Dataset(str(file_path), 'r')
        except ImportError:
            try:
                import xarray as xr
                return self._read_with_xarray(file_path, mode, variable)
            except ImportError:
                return "[警告] 未安装 netCDF4 或 xarray 库。\n请安装: pip install netCDF4 xarray"
        except Exception as e:
            return f"读取NetCDF文件失败: {e}"
        
        result_lines = [f"NetCDF文件: {file_path.name}", f"路径: {file_path}",
                        f"大小: {self._format_size(file_path.stat().st_size)}",
                        f"格式: {dataset.file_format}", ""]
        
        try:
            if mode == "info":
                result_lines.append(f"维度 ({len(dataset.dimensions)}):")
                for name, dim in dataset.dimensions.items():
                    result_lines.append(f"   {name}: {len(dim)}")
                result_lines.extend([f"\n变量数: {len(dataset.variables)}", f"全局属性数: {len(dataset.ncattrs())}"])
                if dataset.ncattrs():
                    result_lines.append("\n全局属性:")
                    for attr_name in dataset.ncattrs():
                        val = str(getattr(dataset, attr_name))[:100]
                        result_lines.append(f"   {attr_name}: {val}")
            elif mode == "vars":
                result_lines.append(f"变量列表 ({len(dataset.variables)}):\n")
                for name, var in dataset.variables.items():
                    dims = ", ".join(var.dimensions) if var.dimensions else "标量"
                    result_lines.append(f"• {name}\n  维度: ({dims})\n  形状: {var.shape}\n  类型: {var.datatype}\n")
            elif mode == "data":
                if variable not in dataset.variables:
                    return f"Error: 变量 '{variable}' 不存在。可用: {', '.join(dataset.variables.keys())}"
                var = dataset.variables[variable]
                total_size = 1
                for s in var.shape:
                    total_size *= s
                if total_size > 10000:
                    sample_slices = tuple(slice(min(s, 10)) for s in var.shape[:3])
                    if len(var.shape) > 3:
                        sample_slices += (0,) * (len(var.shape) - 3)
                    data = var[sample_slices]
                    result_lines.append(f"[警告] 数据过大 ({total_size} 元素)，仅显示样本")
                else:
                    data = var[:]
                import numpy as np
                result_lines.extend([f"变量: {variable}", f"形状: {data.shape}",
                                     f"范围: {data.min():.4f} ~ {data.max():.4f}",
                                     f"平均值: {data.mean():.4f}", f"\n数据样本:\n{str(data)}"])
            elif mode == "stats":
                if variable not in dataset.variables:
                    return f"Error: 变量 '{variable}' 不存在。可用: {', '.join(dataset.variables.keys())}"
                var = dataset.variables[variable]
                data = var[:].flatten()
                import numpy as np
                valid = data[~np.isnan(data)]
                result_lines.extend([f"变量统计: {variable}", f"形状: {var.shape}",
                                     f"   元素总数: {len(data)}", f"   有效值数: {len(valid)}",
                                     f"   最小值: {np.min(valid):.6f}", f"   最大值: {np.max(valid):.6f}",
                                     f"   平均值: {np.mean(valid):.6f}", f"   标准差: {np.std(valid):.6f}"])
            
            dataset.close()
        except Exception as e:
            result_lines.append(f"处理失败: {e}")
        
        return "\n".join(result_lines)
    
    def _read_with_xarray(self, file_path, mode: str, variable: str) -> str:
        try:
            import xarray as xr
            import numpy as np
            ds = xr.open_dataset(file_path)
            lines = [f"NetCDF文件: {file_path.name}", ""]
            if mode == "info":
                lines.extend([f"维度: {list(ds.dims)}", f"变量: {list(ds.data_vars)}", f"坐标: {list(ds.coords)}"])
            elif mode == "vars":
                for name, var in ds.data_vars.items():
                    lines.append(f"• {name}: dims={list(var.dims)}, shape={var.shape}")
            elif mode in ("data", "stats"):
                if variable not in ds.data_vars:
                    return f"Error: 变量 '{variable}' 不存在。可用: {list(ds.data_vars)}"
                var = ds[variable]
                valid = var.values[~np.isnan(var.values)]
                lines.extend([f"变量: {variable}", f"形状: {var.shape}",
                              f"   最小值: {np.min(valid):.6f}", f"   最大值: {np.max(valid):.6f}",
                              f"   平均值: {np.mean(valid):.6f}", f"   标准差: {np.std(valid):.6f}"])
            ds.close()
            return "\n".join(lines)
        except Exception as e:
            return f"xarray 读取失败: {e}"
    
    def _format_size(self, size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"


# ═══════════════════════════════════════════════════════════════════════
#  Enhanced Tool Registry
# ═══════════════════════════════════════════════════════════════════════

class ToolRegistry:
    """Enhanced registry with permission checks and concurrency control."""
    
    def __init__(self):
        self._tools: dict[str, BaseTool] = {}
        self._running_tools: set[str] = set()
        self._lock = threading.Lock()
    
    def register(self, tool: BaseTool):
        self._tools[tool.name] = tool
    
    def get(self, name: str) -> BaseTool | None:
        return self._tools.get(name)
    
    def list_tools(self) -> list[BaseTool]:
        return list(self._tools.values())
    
    def list_definitions(self) -> str:
        """Generate tool definitions for LLM prompt."""
        lines = []
        for tool in self._tools.values():
            params = ", ".join(f"{k}: {v}" for k, v in tool.parameters.items())
            lines.append(f"- **{tool.name}**({params}): {tool.description}")
        return "\n".join(lines)
    
    def check_permission(self, name: str, params: dict) -> tuple[bool, str]:
        """Check if tool execution is allowed."""
        tool = self._tools.get(name)
        if not tool:
            return False, f"Unknown tool: {name}"
        
        perm = tool.check_permission(params)
        if perm == ToolPermission.DENIED:
            return False, f"Tool {name} is denied by policy"
        elif perm == ToolPermission.DANGEROUS:
            return False, f"Tool {name} requires explicit user approval for dangerous operation"
        
        return True, ""
    
    def execute(
        self,
        name: str,
        params: dict,
        ctx: ToolExecutionContext,
        on_progress: Callable[[str], None] | None = None
    ) -> ToolResult:
        """Execute tool with concurrency control and progress reporting."""
        tool = self._tools.get(name)
        if not tool:
            return ToolResult(False, f"Unknown tool: {name}", name)
        
        # Check concurrency
        if not tool.is_concurrent_safe():
            with self._lock:
                if name in self._running_tools:
                    return ToolResult(False, f"Tool {name} is already running", name)
                self._running_tools.add(name)
        
        try:
            # Report progress
            if on_progress:
                on_progress(tool.get_activity_description(params))
            
            start = time.time()
            output = tool.run(params, ctx)
            elapsed = (time.time() - start) * 1000
            
            # Format output
            formatted = tool.format_output(output)
            is_truncated = len(output) > tool.max_result_size
            
            return ToolResult(
                success=True,
                output=formatted,
                tool_name=name,
                elapsed_ms=elapsed,
                is_truncated=is_truncated
            )
        except Exception as e:
            elapsed = (time.time() - start) * 1000
            return ToolResult(
                success=False,
                output=f"{type(e).__name__}: {e}",
                tool_name=name,
                elapsed_ms=elapsed
            )
        finally:
            if not tool.is_concurrent_safe():
                with self._lock:
                    self._running_tools.discard(name)


def build_default_registry() -> ToolRegistry:
    """Build default tool registry with multimodal support."""
    reg = ToolRegistry()
    # Core code tools
    reg.register(SearchTool())
    reg.register(ReadFileTool())
    reg.register(FindPatternTool())
    reg.register(ListDirTool())
    reg.register(WriteFileTool())
    reg.register(SearchReplaceTool())
    reg.register(ShellTool())
    # Multimodal tools
    reg.register(_ImageReaderTool())
    reg.register(_PDFReaderTool())
    reg.register(_DocumentReaderTool())
    reg.register(_FileBrowserTool())
    reg.register(_NCDataReaderTool())
    return reg


# ═══════════════════════════════════════════════════════════════════════
#  Enhanced Memory System
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class MemoryEntry:
    """Single memory entry."""
    role: str          # "user", "agent", "tool", "system"
    content: str
    tool_name: str = ""
    timestamp: float = field(default_factory=time.time)
    token_estimate: int = 0
    
    def __post_init__(self):
        if not self.token_estimate:
            # Rough estimate: 1 token ≈ 4 chars
            self.token_estimate = len(self.content) // 4


class ShortTermMemory:
    """In-session memory with smart context management."""
    
    def __init__(self, max_entries: int = 20, max_tokens: int = 30000):
        self.entries: list[MemoryEntry] = []
        self.max_entries = max_entries
        self.max_tokens = max_tokens
    
    def add(self, role: str, content: str, tool_name: str = ""):
        entry = MemoryEntry(role=role, content=content, tool_name=tool_name)
        self.entries.append(entry)
        
        # Prune if needed
        self._prune()
    
    def _prune(self):
        """Prune memory to stay within limits."""
        # Keep first (goal) and last N entries
        if len(self.entries) > self.max_entries:
            self.entries = [self.entries[0]] + self.entries[-(self.max_entries - 1):]
        
        # Estimate total tokens
        total_tokens = sum(e.token_estimate for e in self.entries)
        while total_tokens > self.max_tokens and len(self.entries) > 2:
            # Remove oldest non-first entry
            removed = self.entries.pop(1)
            total_tokens -= removed.token_estimate
    
    def get_context(self, max_chars: int = 30000) -> str:
        """Format memory into string for LLM."""
        lines = []
        total = 0
        
        for entry in reversed(self.entries):
            content_str = entry.content
            
            # Smart truncation by role
            if entry.role == "tool" and len(content_str) > 2000:
                content_str = content_str[:2000] + "\n...[truncated]..."
            elif entry.role == "agent" and len(content_str) > 1000:
                content_str = content_str[:1000] + "..."
            
            line = f"[{entry.role}] {content_str}"
            if total + len(line) > max_chars:
                break
            lines.append(line)
            total += len(line)
        
        lines.reverse()
        return "\n---\n".join(lines)
    
    def get_recent_tool_results(self, n: int = 3) -> list[str]:
        """Get recent tool results for context."""
        results = []
        for entry in reversed(self.entries):
            if entry.role == "tool":
                results.append(entry.content)
                if len(results) >= n:
                    break
        return list(reversed(results))
    
    def clear(self):
        self.entries.clear()


class LongTermMemory:
    """Persistent memory with enhanced retrieval."""
    
    def __init__(self, project_root: Path):
        self.dir = get_snowcode_dir(project_root)
        self.path = self.dir / "memory.jsonl"
    
    def store(self, question: str, answer: str, actions: list[dict]):
        """Save Q&A session."""
        entry = {
            "ts": time.time(),
            "q": question[:200],
            "a": answer[:500],
            "actions": [a.get("tool", "") for a in actions],
            "hash": hashlib.sha256(question.encode()).hexdigest()[:16],
        }
        
        with open(self.path, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    
    def recall(self, question: str, n: int = 3) -> str:
        """Retrieve similar past Q&As."""
        if not self.path.exists():
            return ""
        
        q_words = set(re.findall(r'\w+', question.lower()))
        q_trigrams = self._get_trigrams(question)
        
        scored = []
        with open(self.path, encoding="utf-8") as f:
            for line in f:
                try:
                    entry = json.loads(line.strip())
                except json.JSONDecodeError:
                    continue
                
                past_q = entry.get("q", "")
                e_words = set(re.findall(r'\w+', past_q.lower()))
                e_trigrams = self._get_trigrams(past_q)
                
                word_overlap = len(q_words & e_words)
                trigram_overlap = len(q_trigrams & e_trigrams)
                
                score = (word_overlap * 5) + trigram_overlap
                if score > 0:
                    scored.append((score, entry))
        
        scored.sort(key=lambda x: -x[0])
        if not scored:
            return ""
        
        lines = []
        for _, e in scored[:n]:
            lines.append(f"Previously asked: \"{e['q']}\"\nAnswer summary: {e['a']}")
        return "\n\n".join(lines)
    
    def _get_trigrams(self, text: str) -> set[str]:
        text = text.lower()
        return {text[i:i+3] for i in range(len(text)-2)} if len(text) >= 3 else set()
    
    def clear(self):
        self.path.unlink(missing_ok=True)


# ═══════════════════════════════════════════════════════════════════════
#  Enhanced LLM Client
# ═══════════════════════════════════════════════════════════════════════

class LLMClient:
    """Enhanced LLM client with caching and retry."""
    
    def __init__(self, model: str | None = None):
        self.model = model
        self.api_key, self.base_url, self.model_name, self.thinking = _get_llm_config(model)
        self._cache: dict[str, str] = {}
    
    @property
    def available(self) -> bool:
        return bool(self.api_key)
    
    def complete(
        self,
        system: str,
        user: str,
        temperature: float = 0.1,
        use_cache: bool = False
    ) -> str:
        """Non-streaming completion with optional caching."""
        if not self.available:
            return ""
        
        # Check cache
        if use_cache:
            cache_key = hashlib.md5(f"{system}:{user}".encode()).hexdigest()
            if cache_key in self._cache:
                return self._cache[cache_key]
        
        result = self._call_llm(system, user, temperature)
        
        # Cache result
        if use_cache and result:
            self._cache[cache_key] = result
        
        return result
    
    def _call_llm(self, system: str, user: str, temperature: float) -> str:
        """Call LLM backend."""
        if self.api_key == "ollama":
            return self._call_ollama(system, user)
        return self._call_openai(system, user, temperature)
    
    def _call_ollama(self, system: str, user: str) -> str:
        """Call Ollama API."""
        try:
            import httpx
            resp = httpx.post(
                f"{self.base_url}/api/chat",
                json={
                    "model": self.model_name,
                    "messages": [
                        {"role": "system", "content": system},
                        {"role": "user", "content": user},
                    ],
                    "stream": False,
                },
                timeout=120,
            )
            if resp.status_code == 200:
                return resp.json()["message"]["content"]
            return f"[LLM Error] HTTP {resp.status_code}: {resp.text}"
        except Exception as e:
            return f"[LLM Error] {type(e).__name__}: {e}"
    
    def _call_openai(self, system: str, user: str, temperature: float) -> str:
        """Call OpenAI-compatible API."""
        try:
            import openai
            client = openai.OpenAI(api_key=self.api_key, base_url=self.base_url)
            extra = {"enable_thinking": self.thinking} if self.thinking else {}
            resp = client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=temperature,
                max_tokens=4096,
                extra_body=extra,
            )
            return resp.choices[0].message.content or ""
        except Exception as e:
            return f"[LLM Error] {type(e).__name__}: {e}"
    
    def analyze_image(self, system: str, user_text: str, image_path: str) -> str:
        """Analyze an image using a multimodal LLM (e.g., qwen-vl-plus)."""
        if not self.available:
            return ""
        try:
            import base64
            import os
            
            # Determine MIME type
            ext = os.path.splitext(image_path)[1].lower()
            mime_map = {
                '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'
            }
            mime = mime_map.get(ext, 'image/jpeg')
            
            # Read and encode image
            with open(image_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode('utf-8')
            data_url = f"data:{mime};base64,{b64}"
            
            # Use OpenAI-compatible client
            import openai
            client = openai.OpenAI(api_key=self.api_key, base_url=self.base_url)
            
            # Use a vision-capable model if available, otherwise fall back to configured model
            # DashScope supports qwen-vl-plus/max for vision
            model_to_use = self.model_name
            if 'vl' not in model_to_use.lower():
                # Suggest a vision model if the current one isn't vision-capable
                # But we'll try the current one first to respect user config
                pass
            
            resp = client.chat.completions.create(
                model=model_to_use,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": [
                        {"type": "text", "text": user_text},
                        {"type": "image_url", "image_url": {"url": data_url}}
                    ]}
                ],
                max_tokens=2048,
            )
            return resp.choices[0].message.content or ""
        except Exception as e:
            return f"[LLM Error] {type(e).__name__}: {e}"


# ═══════════════════════════════════════════════════════════════════════
#  Planning System
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class PlanStep:
    """Single plan step."""
    index: int
    description: str
    tool_hint: str = ""
    status: str = "pending"  # pending / running / done / failed
    result: str = ""


@dataclass
class Plan:
    """Execution plan."""
    goal: str
    steps: list[PlanStep] = field(default_factory=list)
    current: int = 0
    
    @property
    def done(self) -> bool:
        return all(s.status in ("done", "failed") for s in self.steps)
    
    @property
    def current_step(self) -> PlanStep | None:
        for s in self.steps:
            if s.status == "pending":
                return s
        return None
    
    def mark_current(self, status: str, result: str = ""):
        for s in self.steps:
            if s.status == "pending":
                s.status = status
                s.result = result[:200]
                break
    
    def to_context(self) -> str:
        lines = [f"Goal: {self.goal}", ""]
        for s in self.steps:
            icon = {"pending": "[ ]", "running": "[~]", "done": "[+]", "failed": "[-]"}
            lines.append(f"  {icon.get(s.status, '?')} Step {s.index}: {s.description}")
            if s.result:
                lines.append(f"      → {s.result[:100]}")
        return "\n".join(lines)


PLANNER_PROMPT = """\
你是一个任务规划器。根据用户目标，拆解为 2-5 个可执行步骤。

输出 JSON 数组，每个元素：
{{"index": 1, "description": "步骤描述", "tool_hint": "建议使用的工具名(可选)"}}

可用工具：
{tools}

规则：
1. 步骤要具体可执行，不要写"分析代码"这种模糊描述
2. 每个步骤对应一次工具调用
3. 最后一步通常是"综合信息给出回答"
4. 只输出 JSON 数组，不要其他内容
"""


class Planner:
    """Decompose user goals into executable plan steps."""
    
    def __init__(self, llm: LLMClient, tools_desc: str):
        self.llm = llm
        self.tools_desc = tools_desc
    
    def create_plan(self, goal: str) -> Plan:
        prompt = PLANNER_PROMPT.format(tools=self.tools_desc)
        user_msg = f"用户目标: {goal}"
        raw = self.llm.complete(prompt, user_msg)
        
        steps = self._parse_steps(raw)
        if not steps:
            steps = [PlanStep(index=1, description=f"搜索并回答: {goal}", tool_hint="search")]
        
        return Plan(goal=goal, steps=steps)
    
    def refine_plan(self, plan: Plan, observation: str) -> Plan:
        """Adjust remaining steps based on new information."""
        failed = [s for s in plan.steps if s.status == "failed"]
        for s in failed:
            s.status = "pending"
            s.tool_hint = "find_pattern" if s.tool_hint == "search" else "search"
        return plan
    
    def _parse_steps(self, raw: str) -> list[PlanStep]:
        raw = raw.strip()
        
        # Try JSON blocks
        json_blocks = re.findall(r"```(?:json)?\s*(.*?)\s*```", raw, re.DOTALL)
        for block in json_blocks:
            try:
                data = json.loads(block)
                if isinstance(data, list):
                    return [
                        PlanStep(
                            index=item.get("index", i+1),
                            description=item.get("description", ""),
                            tool_hint=item.get("tool_hint", ""),
                        )
                        for i, item in enumerate(data)
                        if isinstance(item, dict)
                    ]
            except Exception:
                continue
        
        # Try direct JSON
        try:
            data = json.loads(raw)
            if isinstance(data, list):
                return [
                    PlanStep(
                        index=item.get("index", i+1),
                        description=item.get("description", ""),
                        tool_hint=item.get("tool_hint", ""),
                    )
                    for i, item in enumerate(data)
                    if isinstance(item, dict)
                ]
        except Exception:
            pass
        
        return []


# ═══════════════════════════════════════════════════════════════════════
#  Agent Result
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class AgentResult:
    """Enhanced agent result with detailed metrics."""
    answer: str
    plan: Plan | None
    actions: list[dict]
    steps_taken: int
    memory_entries: int
    total_tokens_used: int = 0
    total_elapsed_ms: float = 0
    tools_used: list[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════
#  Enhanced CodeAgent
# ═══════════════════════════════════════════════════════════════════════

AGENT_SYSTEM = """\
你是一个强大的代码 Agent，通过工具查找、分析并修改代码。

当前操作系统: {platform}
使用 shell 工具时请使用对应平台的命令语法（Windows用cmd命令，Linux/Mac用bash命令）。

## 输出格式（每次回复必须是 JSON）

调用工具：
{{"think": "思考过程", "tool": "工具名", "params": {{"参数": "值"}}}}

给出结论：
{{"think": "思考过程", "answer": "最终回答(Markdown)"}}

## 规则

1. 每次只调一个工具
2. 不要重复搜索相同关键词
3. 读文件时**一次读完整个文件**（不要分段读），除非文件超过2000行
4. 连续 3 次无结果则直接回答
5. 最多 {max_steps} 轮
6. 回答直接了当，不要"根据上下文"等废话
7. 精确标注代码位置(文件:行号)
8. 中文回答，代码术语英文
9. 当被要求修改代码、修复 Bug 或生成测试时，必须使用 write_file 或 search_replace 工具直接将代码写入项目
10. 当被要求生成文件、报告、文档、图表时，必须使用 write_file 工具将内容写入项目，不要只在回答中显示！
11. 写入文件时，path 参数使用相对路径（相对于项目根目录）
"""


class CodeAgent:
    """
    Enhanced ReAct Agent with Claude Code-inspired features.
    
    Improvements over v1:
    1. Tool System: Permission checks, concurrency control, progress reporting
    2. Memory: Smart context management, token estimation
    3. Planning: Better plan refinement
    4. Execution: Better error handling, repeat detection
    """
    
    def __init__(
        self,
        store: VectorStore,
        project_root: Path,
        model: str | None = None,
        max_steps: int = 5,
        use_planning: bool = True,
    ):
        self.store = store
        self.root = project_root.resolve()
        self.max_steps = max_steps
        self.use_planning = use_planning
        
        # Components
        self.llm = LLMClient(model)
        self.tools = build_default_registry()
        self.memory_st = ShortTermMemory(max_entries=20, max_tokens=30000)
        self.memory_lt = LongTermMemory(self.root)
        self.planner = Planner(self.llm, self.tools.list_definitions())
        
        self.ctx = ToolExecutionContext(
            root=self.root,
            store=store,
            llm=self.llm
        )
    
    def run(
        self,
        question: str,
        on_step: Callable[[int, str, str], None] | None = None,
        on_think: Callable[[str], None] | None = None,
        on_answer: Callable[[str], None] | None = None,
        on_progress: Callable[[str], None] | None = None,
    ) -> AgentResult:
        """Run the agent loop with enhanced features."""
        start_time = time.time()
        actions_log = []
        tools_used = set()
        
        # 1. Recall long-term memory
        past = self.memory_lt.recall(question)
        if past:
            self.memory_st.add("system", f"相关历史记忆:\n{past}")
        
        # 2. Create plan
        plan = None
        if self.use_planning:
            plan = self.planner.create_plan(question)
            if plan and on_progress:
                on_progress(f"  计划: {plan.goal}")
                for s in plan.steps:
                    on_progress(f"    - 步骤 {s.index}: {s.description} [{s.tool_hint or '通用'}]")
                on_progress("")
        
        # 3. Execute loop
        answer = ""
        no_result_streak = 0
        json_retries = 0
        MAX_JSON_RETRIES = 3
        
        step = 0
        max_steps = self.max_steps if self.max_steps > 0 else 50
        action_history: list[tuple[str, str]] = []
        repeat_count = 0
        
        while step < max_steps:
            # Check abort signal
            if self.ctx.abort_signal and self.ctx.abort_signal.is_set():
                answer = "Operation cancelled by user."
                break
            
            # Build prompt
            system = AGENT_SYSTEM.format(max_steps=max_steps, platform=sys.platform)
            tools_desc = self.tools.list_definitions()
            mem_ctx = self.memory_st.get_context()
            
            plan_ctx = ""
            if plan:
                plan_ctx = f"\n## 当前计划\n{plan.to_context()}\n"
            
            user_msg = f"""## 工具列表
{tools_desc}
{plan_ctx}
## 记忆
{mem_ctx}

## 用户问题
{question}

请思考下一步。输出 JSON:"""
            
            raw = self.llm.complete(system, user_msg, temperature=0.2)
            
            if not raw:
                # Fallback to direct search
                results = self.store.query(question, n_results=5)
                if results:
                    from .rag import _format_context
                    answer = "LLM 不可用，相关代码：\n\n" + _format_context(results)
                else:
                    answer = "未找到相关代码，LLM 不可用。"
                break
            
            # Parse JSON response
            try:
                parsed = self._parse_json(raw)
                json_retries = 0
            except ValueError as e:
                json_retries += 1
                if json_retries >= MAX_JSON_RETRIES:
                    answer = f"Error: LLM consistently returns invalid JSON. Raw:\n{raw[:500]}"
                    break
                
                err_msg = f"System Error: {e}. Please output valid JSON only."
                self.memory_st.add("system", err_msg)
                if on_step:
                    on_step(step + 1, "system_error", "JSON parse failed, retrying...")
                continue
            
            step += 1
            think = parsed.get("think", "")
            
            if on_think and think:
                on_think(think)
            self.memory_st.add("agent", f"Think: {think}")
            
            # Check for final answer
            if "answer" in parsed and "tool" not in parsed:
                answer = parsed["answer"]
                break
            
            # Execute tool
            tool_name = parsed.get("tool", "")
            params = parsed.get("params", {})
            
            if not tool_name:
                answer = parsed.get("answer", raw)
                break
            
            # Check permission
            allowed, perm_msg = self.tools.check_permission(tool_name, params)
            if not allowed:
                self.memory_st.add("system", f"Permission denied: {perm_msg}")
                if on_step:
                    on_step(step, tool_name, f"Permission denied: {perm_msg}")
                continue
            
            # Execute tool with progress reporting
            result = self.tools.execute(
                tool_name, params, self.ctx,
                on_progress=on_progress
            )
            
            # Log action
            actions_log.append({
                "tool": tool_name,
                "params": {k: str(v)[:50] for k, v in params.items()},
                "success": result.success,
                "elapsed_ms": result.elapsed_ms,
            })
            tools_used.add(tool_name)
            
            if on_step:
                on_step(step, tool_name, result.preview)
            
            # Store in memory
            self.memory_st.add("tool", f"[{tool_name}] {result.output[:8000]}", tool_name=tool_name)
            
            # Repeat detection
            action_key = (tool_name, json.dumps(params, sort_keys=True, ensure_ascii=False)[:100])
            if action_key in action_history:
                repeat_count += 1
            else:
                repeat_count = 0
            action_history.append(action_key)
            
            if repeat_count >= 2:
                answer = f"检测到重复操作（连续 {repeat_count+1} 次），强制输出结论：\n\n"
                summary_prompt = f"基于已收集的信息，直接回答用户问题：{question}"
                summary = self.llm.complete("直接回答，不要调用工具。", summary_prompt, temperature=0.3)
                answer += summary if summary else "无法得出结论。"
                break
            
            # Update plan
            if plan:
                if result.success:
                    plan.mark_current("done", result.output[:100])
                else:
                    plan.mark_current("failed", result.output[:100])
                    self.planner.refine_plan(plan, result.output)
            
            # Track no-result streak
            if result.success and len(result.output) < 20:
                no_result_streak += 1
            else:
                no_result_streak = 0
            
            if no_result_streak >= 3:
                answer = "连续多次未找到有效信息，请尝试更具体的问题。"
                break
        
        if not answer:
            answer = "已达到最大步数。"
        
        # 4. Store to long-term memory
        self.memory_lt.store(question, answer, actions_log)
        
        if on_answer:
            on_answer(answer)
        
        elapsed = (time.time() - start_time) * 1000
        
        return AgentResult(
            answer=answer,
            plan=plan,
            actions=actions_log,
            steps_taken=len(actions_log),
            memory_entries=len(self.memory_st.entries),
            total_elapsed_ms=elapsed,
            tools_used=list(tools_used),
        )
    
    def _parse_json(self, raw: str) -> dict | list:
        """Parse JSON robustly from LLM output."""
        raw = raw.strip()
        
        # Try markdown JSON blocks
        json_blocks = re.findall(r"```(?:json)?\s*(.*?)\s*```", raw, re.DOTALL)
        for block in json_blocks:
            try:
                return json.loads(block)
            except json.JSONDecodeError:
                continue
        
        # Try finding balanced braces
        for start_char, end_char in [("{", "}"), ("[", "]")]:
            start = raw.find(start_char)
            if start == -1:
                continue
            depth = 0
            for i in range(start, len(raw)):
                if raw[i] == start_char:
                    depth += 1
                elif raw[i] == end_char:
                    depth -= 1
                if depth == 0:
                    try:
                        return json.loads(raw[start:i+1])
                    except json.JSONDecodeError:
                        break
        
        # Try direct parse
        try:
            return json.loads(raw)
        except Exception:
            pass
        
        # Extract answer from plain text
        if "answer" in raw.lower() or "答" in raw:
            for marker in ["Answer:", "answer:", "回答：", "答案："]:
                idx = raw.find(marker)
                if idx != -1:
                    answer_text = raw[idx + len(marker):].strip()
                    return {"think": "", "answer": answer_text}
        
        return {"think": "", "answer": raw}
    
    def reset_memory(self):
        """Clear all memory."""
        self.memory_st.clear()
        self.memory_lt.clear()


# ═══════════════════════════════════════════════════════════════════════
#  Multi-Agent Coordinator (Inspired by Claude Code)
# ═══════════════════════════════════════════════════════════════════════

class AgentRole(Enum):
    """Agent roles in multi-agent system."""
    COORDINATOR = "coordinator"  # Plans and delegates
    WORKER = "worker"           # Executes tasks
    VERIFIER = "verifier"       # Verifies results


@dataclass
class WorkerTask:
    """Task for a worker agent."""
    task_id: str
    description: str
    prompt: str
    status: str = "pending"  # pending, running, completed, failed
    result: str = ""


class MultiAgentCoordinator:
    """
    Multi-agent coordination system inspired by Claude Code's Coordinator mode.
    
    The Coordinator:
    - Understands user goals
    - Decomposes tasks
    - Delegates to Workers
    - Synthesizes results
    """
    
    def __init__(
        self,
        store: VectorStore,
        project_root: Path,
        model: str | None = None,
        num_workers: int = 2,
    ):
        self.store = store
        self.root = project_root
        self.model = model
        self.num_workers = num_workers
        
        self.llm = LLMClient(model)
        self.tasks: dict[str, WorkerTask] = {}
        self.results: dict[str, str] = {}
        self.on_progress: Callable[[str], None] | None = None
    
    def coordinate(
        self,
        question: str,
        on_progress: Callable[[str], None] | None = None,
    ) -> str:
        """
        Coordinate multiple workers to answer a complex question.
        """
        # Store on_progress for worker callbacks
        self.on_progress = on_progress
        
        # Step 1: Analyze and decompose
        if on_progress:
            on_progress("Analyzing question...")
        
        subtasks = self._decompose_task(question)
        
        if not subtasks:
            # Fallback to single agent
            if on_progress:
                on_progress("Task decomposition returned empty, falling back to single agent...")
            agent = CodeAgent(self.store, self.root, self.model)
            result = agent.run(question)
            if on_progress:
                on_progress(f"Single agent completed. Steps: {result.steps_taken}")
            return result.answer
        
        if on_progress:
            for i, t in enumerate(subtasks, 1):
                on_progress(f"  Subtask {i}: {t.description}")
        
        # Step 2: Execute workers in parallel
        if on_progress:
            on_progress(f"Executing {len(subtasks)} subtasks in parallel...")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.num_workers) as executor:
            futures = {}
            for task in subtasks:
                if on_progress:
                    on_progress(f"  [Worker] Starting: {task.description}")
                future = executor.submit(self._execute_worker_task, task)
                futures[future] = task
            
            for future in concurrent.futures.as_completed(futures):
                task = futures[future]
                try:
                    result = future.result()
                    self.results[task.task_id] = result
                    task.status = "completed"
                    task.result = result[:200]
                    if on_progress:
                        on_progress(f"  [Done] {task.description}")
                except Exception as e:
                    task.status = "failed"
                    task.result = str(e)
                    if on_progress:
                        on_progress(f"  [Failed] {task.description}: {e}")
        
        # Step 3: Synthesize results
        if on_progress:
            on_progress("Synthesizing results...")
        
        return self._synthesize_results(question)
    
    def _decompose_task(self, question: str) -> list[WorkerTask]:
        """Decompose question into subtasks."""
        prompt = """\
你是一个任务分解专家。将用户的复杂问题分解为 2-4 个可并行执行的子任务。

输出 JSON 数组：
[{"task_id": "1", "description": "子任务描述", "prompt": "给worker的完整指令"}]

规则：
1. 每个子任务应该是独立的，可以并行执行
2. prompt 要完整自包含，worker看不到原始问题
3. 只输出 JSON 数组
"""
        
        raw = self.llm.complete(prompt, f"用户问题: {question}")
        
        if self.on_progress:
            self.on_progress(f"  LLM raw response (first 200 chars): {raw[:200]}")
        
        try:
            # Parse JSON
            json_blocks = re.findall(r"```(?:json)?\s*(.*?)\s*```", raw, re.DOTALL)
            for block in json_blocks:
                try:
                    data = json.loads(block)
                    if isinstance(data, list):
                        return [
                            WorkerTask(
                                task_id=item.get("task_id", str(i+1)),
                                description=item.get("description", ""),
                                prompt=item.get("prompt", ""),
                            )
                            for i, item in enumerate(data)
                            if isinstance(item, dict)
                        ]
                except Exception:
                    continue
            
            data = json.loads(raw)
            if isinstance(data, list):
                return [
                    WorkerTask(
                        task_id=item.get("task_id", str(i+1)),
                        description=item.get("description", ""),
                        prompt=item.get("prompt", ""),
                    )
                    for i, item in enumerate(data)
                    if isinstance(item, dict)
                ]
        except Exception:
            pass
        
        return []
    
    def _execute_worker_task(self, task: WorkerTask) -> str:
        """Execute a single worker task."""
        task.status = "running"
        
        # Create a worker agent
        worker = CodeAgent(
            self.store,
            self.root,
            self.model,
            max_steps=10,
            use_planning=False,
        )
        
        def on_worker_step(num, tool_name, preview):
            if self.on_progress:
                preview_str = preview[:80] + "..." if len(preview) > 80 else preview
                self.on_progress(f"    [Worker {task.task_id}] Step {num}: {tool_name} -> {preview_str}")
        
        result = worker.run(task.prompt, on_step=on_worker_step)
        return result.answer
    
    def _synthesize_results(self, original_question: str) -> str:
        """Synthesize results from all workers."""
        if not self.results:
            return "No results from workers."
        
        results_text = "\n\n".join([
            f"Subtask {task_id}: {result}"
            for task_id, result in self.results.items()
        ])
        
        prompt = f"""\
你是一个结果综合专家。根据多个子任务的结果，综合回答用户的原始问题。

原始问题: {original_question}

子任务结果:
{results_text}

请综合以上信息，给出完整、准确的回答。"""
        
        return self.llm.complete("直接回答，不要调用工具。", prompt, temperature=0.3)


# ═══════════════════════════════════════════════════════════════════════
#  Convenience Functions
# ═══════════════════════════════════════════════════════════════════════

def create_agent(
    store: VectorStore,
    project_root: Path,
    model: str | None = None,
    **kwargs
) -> CodeAgent:
    """Create an enhanced CodeAgent."""
    return CodeAgent(store, project_root, model, **kwargs)


def create_coordinator(
    store: VectorStore,
    project_root: Path,
    model: str | None = None,
    **kwargs
) -> MultiAgentCoordinator:
    """Create a multi-agent coordinator."""
    return MultiAgentCoordinator(store, project_root, model, **kwargs)
